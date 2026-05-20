from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
import numpy as np
import re
import openpyxl
import io
import os

# 1. Obtener la ruta absoluta de la carpeta donde está app.py (backend/)
base_dir = os.path.abspath(os.path.dirname(__file__))

# 2. Construir las rutas hacia la carpeta frontend
# ".." significa "subir un nivel" (salir de backend y entrar a frontend)
template_dir = os.path.join(base_dir, '../frontend/templates')
static_dir = os.path.join(base_dir, '../frontend')

# 3. Inicializar Flask con las rutas correctas
app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)

CORS(app)


def parse_excel(file_bytes):
    """
    Lee un archivo binario de Excel (.xlsx o .xlsm) en memoria,
    extrayendo las dimensiones y los vectores tabulares del ensayo.
    """
    # Cargar el libro desde los bytes en memoria (data_only=True extrae valores, no fórmulas)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active  # Toma la hoja por defecto o usa wb["NombreDeLaHoja"]

    # 1. Extraer Metadatos 
    area = sheet['D1'].value
    length = sheet['D2'].value

    # Forzar la conversión a flotante por si vienen como texto
    try: area = float(area) if area else None
    except ValueError: area = None

    try: length = float(length) if length else None
    except ValueError: length = None

    # 2. Extraer Datos Tabulares (Load y Stroke)
    load_list = []
    stroke_list = []

    fila_inicio = 2
    for row in range(fila_inicio, sheet.max_row + 1):
        load_cell = sheet.cell(row=row, column=2).value
        stroke_cell = sheet.cell(row=row, column=1).value

        # Romper el ciclo si encontramos celdas vacías al final de la tabla
        if load_cell is None or stroke_cell is None:
            continue

        try:
            load_list.append(float(load_cell))
            stroke_list.append(float(stroke_cell))
        except ValueError:
            # Ignora filas con texto o encabezados intermedios
            continue

    # Normalizar restando el valor inicial (idéntico al pipeline de archivos .txt)
    if load_list:
        load_list = [l / 9.81 for l in load_list]


    return area, length, load_list, stroke_list

def parse_document(text):
    """Extrae el área, la longitud y las columnas de datos crudos del archivo txt"""
    
    # A. Extraer metadatos (A0, L0)
    area_match = re.search(r'Area[\s:=]+([\d,.]+)', text, re.IGNORECASE)
    length_match = re.search(r'OriginalGL[\s:=]+([\d,.]+)', text, re.IGNORECASE)

    area = float(area_match.group(1).replace(',', '.')) if area_match else None
    length = float(length_match.group(1).replace(',', '.')) if length_match else None

    # B. Procesar datos tabulares
    lines = text.split('\n')
    is_data_section = False
    load_list = []
    stroke_list = []

    for line in lines:
        line = line.strip()
        if not line: continue

        # Identificar cuándo inician los datos (ignorando encabezados)
        if not is_data_section and re.search(r'Load|Stroke|Elongation', line, re.IGNORECASE):
            continue

        # Validar línea estrictamente numérica
        if re.match(r'^[\d,.-]+[\s,;]+[\d,.-]+', line):
            is_data_section = True
            
            columns = re.split(r'[\s;]+', line)
            if len(columns) >= 3:
                try:
                    # Se parsean Load y Stroke asumiendo columnas 1 y 2
                    load_val = float(columns[1].replace(',', '.'))
                    stroke_val = float(columns[2].replace(',', '.'))
                    load_list.append(load_val)
                    stroke_list.append(stroke_val)
                except ValueError:
                    continue

    # Cortar el primer elemento como en la lógica original
    if len(load_list) > 1:
        load_list = load_list[1:]
        stroke_list = stroke_list[1:]

    # Normalizar restando el valor inicial (para arrancar en cero)
    if load_list and stroke_list:
        first_load = load_list[0]
        first_stroke = stroke_list[0]
        load_list = [l - first_load for l in load_list]
        stroke_list = [s - first_stroke for s in stroke_list]

    return area, length, load_list, stroke_list

# Actualiza la firma de la función para recibir los rangos
def procesar_ensayo(load, stroke, A, L0, units, range_min=10.0, range_max=40.0):
    load = np.array(load)
    stroke = np.array(stroke)
    
    g = 9.81 if units == 'Metrico' else 32.174
    factor_fuerza = g * 1000 
    
    force = load * factor_fuerza
    stress = force / A
    strain = stroke / L0

    # --- LÓGICA DINÁMICA DE RANGOS ---
    # Encontrar el Esfuerzo Máximo (UTS)
    UTS = np.max(stress)
    
    # Calcular los límites en MPa según el porcentaje enviado por el usuario
    limite_inferior = UTS * (range_min / 100.0)
    limite_superior = UTS * (range_max / 100.0)

    # np.argmax retorna el primer índice donde la condición se cumple (True)
    idx_inicio = np.argmax(stress >= limite_inferior)
    idx_fin = np.argmax(stress >= limite_superior)
    
    # Prevenir que los índices choquen en curvas irregulares
    if idx_inicio >= idx_fin:
        idx_fin = idx_inicio + 1

    # Regresión lineal usando la porción exacta solicitada por el usuario
    m, b_lin = np.polyfit(strain[idx_inicio:idx_fin], stress[idx_inicio:idx_fin], 1)
    epsilon_0 = -b_lin / m
    strain_corrected = strain - epsilon_0

    y_offset_line = m * (strain_corrected - 0.002)
    diferencia = stress - y_offset_line
    
    idx_fluencia = np.where(diferencia[idx_fin:] < 0)[0]
    if len(idx_fluencia) > 0:
        idx_real = idx_fluencia[0] + idx_fin
    else:
        idx_real = np.argmin(np.abs(diferencia[idx_fin:])) + idx_fin
        
    Sy = stress[idx_real]
    x_Sy = strain_corrected[idx_real]
    
    ## Linea del modulo de Young
    y_plot_elastic = np.linspace(0, Sy, 100)
    x_plot_elastic = y_plot_elastic / m

    ## Listas de variables para graficar
    strain_position = np.where(strain_corrected >= 0)
    strain_graff = strain_corrected[strain_position]
    stress_graff = stress[strain_position] 
    
    return {
        "E": round(m / 1000, 2),
        "Sy": round(Sy, 2),
        "Sut": round(UTS, 2),
        "Elong": round(np.max(strain_corrected) * 100,2),
        "strain_corrected": strain_graff.tolist(),
        "strain": strain.tolist(),
        "stress": stress.tolist(),
        "stress_corrected": stress_graff.tolist(),
        "x_Sy": x_Sy,
        "y_Sy": Sy,
        "x_plot_elastic": x_plot_elastic.tolist(),
        "y_plot_elastic": y_plot_elastic.tolist()
    }

@app.route('/upload-and-calculate', methods=['POST'])
def upload_and_calculate():
    if 'file' not in request.files:
        return jsonify({"error": "No se proporcionó ningún archivo"}), 400
        
    file = request.files['file']
    filename = file.filename.lower()
    units = request.form.get('units', 'Metrico')


    # --- ENRUTAMIENTO INTELIGENTE SEGÚN EXTENSIÓN ---
    if filename.endswith(('.xlsx', '.xlsm')):
        # 1. Lectura BINARIA para archivos de Excel (No se decodifica como texto)
        file_bytes = file.read() 
        parsed_area, parsed_length, load, stroke = parse_excel(file_bytes)
    else:
        # 2. Lectura de TEXTO exclusivamente para .txt o .csv
        text = file.read().decode('utf-8')
        parsed_area, parsed_length, load, stroke = parse_document(text)
    
    # Preferir dimensiones ingresadas manualmente por el usuario en pantalla si existen
    area = float(request.form.get('area')) if request.form.get('area') else parsed_area
    length = float(request.form.get('length')) if request.form.get('length') else parsed_length

    range_min = float(request.form.get('rangeMin', 10.0))
    range_max = float(request.form.get('rangeMax', 40.0))

    if not area or not length:
        return jsonify({"error": "Falta el área o la longitud de la probeta", "area": area, "length": length}), 400

    if not load or not stroke:
        return jsonify({"error": "No se pudieron extraer datos tabulares válidos del archivo"}), 400

    try:
        # El motor de cálculo matemático sigue intacto y unificado
        results = procesar_ensayo(load, stroke, area, length, units, range_min, range_max)
        results['parsed_area'] = area
        results['parsed_length'] = length
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    
@app.route('/')
def portafolio():
    # Renderiza la página de inicio (Landing Page)
    return render_template('portafolio.html')

@app.route('/tensile-test')
def tensile_test():
    # Renderiza la herramienta de tracción
    return render_template('tensileTest.html')

if __name__ == '__main__':
    app.run(debug=True, port=5000)